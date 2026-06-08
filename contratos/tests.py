# Criado por José Eduardo Santana Martins e OpenAI Codex em 06/06/2026
# Objetivo: Validar cálculos críticos, ACL básica e exportação do módulo de contratos.

from datetime import date
from decimal import Decimal
from io import BytesIO
import json
from django.core.files.uploadedfile import SimpleUploadedFile

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from acls.models import Recurso, RegraAcesso

from .forms import ContratoForm
from .models import (
    AvaliacaoCriterioCompetencia,
    AvaliacaoQualidadeCompetencia,
    ChecklistPagamentoModelo,
    CompetenciaPagamento,
    Contrato,
    ContratoDetalhamentoItem,
    ContratoItem,
    CriterioAvaliacaoQualidade,
    EmpresaContratada,
    EventoFinanceiroContrato,
    EventoFinanceiroItem,
    GrupoAvaliacaoQualidade,
    MedicaoItemCompetencia,
    ModeloAvaliacaoQualidade,
    OcorrenciaContrato,
    TermoAditivo,
)


class ContratosBaseTest(TestCase):
    """Base compartilhada para instanciar usuários, ACL e contrato de teste."""

    def setUp(self):
        User = get_user_model()
        self.admin = User.objects.create_user(username='admin-contratos', password='123', is_staff=True)
        self.user = User.objects.create_user(username='joao', password='123')
        self.sem_acesso = User.objects.create_user(username='maria', password='123')
        self.recurso, _ = Recurso.objects.get_or_create(slug='contratos', defaults={'nome': 'Contratos'})
        RegraAcesso.objects.create(recurso=self.recurso, usuario=self.user, nivel=RegraAcesso.NIVEL_MODIFICACAO)
        self.empresa = EmpresaContratada.objects.create(razao_social='Empresa XPTO', cnpj='00.000.000/0001-00')
        self.contrato = Contrato.objects.create(
            numero_contrato='01/2026',
            apelido='Limpeza predial',
            objeto='Prestação de serviços de limpeza',
            detalhamento_objeto='Serviço contínuo.',
            data_inicio_vigencia=date(2026, 1, 1),
            prazo_inicial_meses=12,
            vigencia_maxima_meses=24,
            empresa_contratada=self.empresa,
            fiscal_administrativo=self.admin,
            fiscal_tecnico=self.admin,
            gestor_contrato=self.admin,
            base_mensal=Decimal('1000.00'),
            criado_por=self.admin,
            atualizado_por=self.admin,
        )


class ContratoCalculoTests(ContratosBaseTest):
    """Cobre cálculos estruturais de contrato, avaliação e retroatividade."""

    def test_competencias_nascem_bloqueadas_sem_checklist_padrao(self):
        competencia = self.contrato.competencias.get(periodo_inicio=date(2026, 1, 1), periodo_fim=date(2026, 1, 31))

        self.assertEqual(competencia.status, CompetenciaPagamento.Status.BLOQUEADO)
        self.assertTrue(competencia.aguardando_checklist_padrao)
        self.assertFalse(competencia.pode_liberar)

    def test_gera_competencias_automaticas_com_base_mensal_em_todos_os_periodos(self):
        contrato = Contrato.objects.create(
            numero_contrato='002/2026',
            apelido='Contrato parcial',
            objeto='Serviço mensal',
            detalhamento_objeto='Detalhamento parcial.',
            data_inicio_vigencia=date(2026, 1, 25),
            prazo_inicial_meses=1,
            vigencia_maxima_meses=12,
            empresa_contratada=self.empresa,
            fiscal_administrativo=self.admin,
            fiscal_tecnico=self.admin,
            gestor_contrato=self.admin,
            base_mensal=Decimal('3100.00'),
            criado_por=self.admin,
            atualizado_por=self.admin,
        )

        competencias = list(contrato.competencias.order_by('periodo_inicio'))
        self.assertEqual(len(competencias), 2)
        self.assertEqual(competencias[0].periodo_inicio, date(2026, 1, 25))
        self.assertEqual(competencias[0].periodo_fim, date(2026, 1, 31))
        self.assertEqual(competencias[0].valor_previsto, Decimal('3100.00'))
        self.assertEqual(competencias[1].periodo_inicio, date(2026, 2, 1))
        self.assertEqual(competencias[1].periodo_fim, date(2026, 2, 24))
        self.assertEqual(competencias[1].valor_previsto, Decimal('3100.00'))

    def test_cadastro_do_checklist_padrao_replicado_desbloqueia_competencias(self):
        ChecklistPagamentoModelo.objects.create(contrato=self.contrato, ordem=1, titulo='FGTS', obrigatorio=True)

        competencias = list(self.contrato.competencias.order_by('periodo_inicio'))
        self.assertTrue(all(competencia.checklist_itens.count() == 1 for competencia in competencias))
        self.assertTrue(all(not competencia.aguardando_checklist_padrao for competencia in competencias))
        self.assertEqual(competencias[0].status, CompetenciaPagamento.Status.RASCUNHO)

    def test_reordena_checklist_padrao_existente_em_sequencia(self):
        ChecklistPagamentoModelo.objects.create(contrato=self.contrato, ordem=1, titulo='FGTS', obrigatorio=True)
        item = ChecklistPagamentoModelo.objects.create(contrato=self.contrato, ordem=1, titulo='Teste 02', obrigatorio=True)

        itens = list(self.contrato.checklist_modelos.order_by('ordem', 'id').values_list('titulo', 'ordem'))
        self.assertEqual(itens, [('FGTS', 1), ('Teste 02', 2)])
        self.assertEqual(item.ordem, 2)

    def test_prorrogacao_cria_competencias_automaticas_adicionais(self):
        TermoAditivo.objects.create(
            contrato=self.contrato,
            numero_termo='1º TA',
            tipo=TermoAditivo.Tipo.PRORROGACAO,
            data_assinatura=date(2026, 11, 1),
            data_inicio=date(2027, 1, 1),
            data_termino=date(2027, 12, 31),
            quantidade_meses=12,
        )

        self.assertEqual(self.contrato.competencias.count(), 24)
        self.assertTrue(
            self.contrato.competencias.filter(
                periodo_inicio=date(2027, 12, 1),
                periodo_fim=date(2027, 12, 31),
            ).exists()
        )

    def test_calcula_subtotal_e_valor_global(self):
        ContratoItem.objects.create(
            contrato=self.contrato,
            ordem=1,
            descricao='Posto A',
            quantidade=Decimal('2.00'),
            valor_unitario=Decimal('150.00'),
        )
        ContratoItem.objects.create(
            contrato=self.contrato,
            ordem=2,
            descricao='Posto B',
            quantidade=Decimal('1.00'),
            valor_unitario=Decimal('100.00'),
        )

        self.contrato.refresh_from_db()
        self.assertEqual(self.contrato.base_mensal, Decimal('400.00'))
        self.assertEqual(self.contrato.valor_global, Decimal('4800.00'))

    def test_calcula_vigencia_periodo_e_regime(self):
        TermoAditivo.objects.create(
            contrato=self.contrato,
            numero_termo='1º TA',
            tipo=TermoAditivo.Tipo.PRORROGACAO,
            data_assinatura=date(2026, 11, 1),
            data_inicio=date(2027, 1, 1),
            data_termino=date(2027, 12, 31),
            quantidade_meses=12,
        )
        self.assertEqual(self.contrato.regime_atual, Contrato.Regime.ORDINARIO)
        self.assertEqual(self.contrato.data_limite_vigencia.strftime('%d/%m/%Y'), '31/12/2027')
        self.assertIn('/24 meses', self.contrato.periodo_acumulado_display)

    def test_bloqueia_pagamento_com_checklist_incompleto(self):
        ChecklistPagamentoModelo.objects.create(contrato=self.contrato, ordem=1, titulo='NF', obrigatorio=True)
        competencia = self.contrato.competencias.get(periodo_inicio=date(2026, 2, 1), periodo_fim=date(2026, 2, 28))
        self.assertFalse(competencia.pode_liberar)

    def test_criacao_e_calculo_de_medicao(self):
        item = ContratoItem.objects.create(
            contrato=self.contrato,
            ordem=1,
            descricao='Item medido',
            quantidade=Decimal('10.00'),
            valor_unitario=Decimal('50.00'),
        )
        competencia = self.contrato.competencias.get(periodo_inicio=date(2026, 1, 1), periodo_fim=date(2026, 1, 31))
        MedicaoItemCompetencia.objects.create(
            competencia=competencia,
            item_contrato=item,
            quantidade=Decimal('2.00'),
            valor_unitario_aplicado=Decimal('50.00'),
        )
        competencia.refresh_from_db()
        self.assertEqual(competencia.valor_medido, Decimal('100.00'))
        self.assertEqual(competencia.valor_liberado, Decimal('100.00'))

    def test_item_do_contrato_usa_proxima_ordem_quando_campo_fica_em_branco(self):
        ContratoItem.objects.create(
            contrato=self.contrato,
            ordem=1,
            descricao='Item inicial',
            quantidade=Decimal('1.00'),
            valor_unitario=Decimal('10.00'),
        )
        self.client.login(username='admin-contratos', password='123')

        response = self.client.post(
            reverse('contratos:item_create', args=[self.contrato.pk]),
            {
                'ordem': '',
                'descricao': 'Item automático',
                'codigo_siafisico': '',
                'codigo_catmat_catser': '',
                'unidade_fornecimento': '',
                'quantidade': '2.00',
                'valor_unitario': '15.00',
                'valor_referencial': '0.00',
            },
        )

        item = ContratoItem.objects.get(descricao='Item automático')
        self.assertRedirects(response, reverse('contratos:contrato_detail', args=[self.contrato.pk]), fetch_redirect_response=False)
        self.assertEqual(item.ordem, 2)

    def test_aplica_avaliacao_qualidade_ao_valor_final(self):
        item = ContratoItem.objects.create(
            contrato=self.contrato,
            ordem=1,
            descricao='Item qualidade',
            quantidade=Decimal('1.00'),
            valor_unitario=Decimal('100.00'),
        )
        competencia = self.contrato.competencias.get(periodo_inicio=date(2026, 3, 1), periodo_fim=date(2026, 3, 31))
        MedicaoItemCompetencia.objects.create(
            competencia=competencia,
            item_contrato=item,
            quantidade=Decimal('1.00'),
            valor_unitario_aplicado=Decimal('100.00'),
        )
        modelo = ModeloAvaliacaoQualidade.objects.create(
            contrato=self.contrato,
            nome='SLA',
            vigencia_inicio=date(2026, 1, 1),
        )
        grupo = GrupoAvaliacaoQualidade.objects.create(modelo=modelo, ordem=1, nome='Desempenho', peso=Decimal('1.00'))
        criterio = CriterioAvaliacaoQualidade.objects.create(
            grupo=grupo,
            ordem=1,
            nome='Pontualidade',
            peso=Decimal('1.00'),
            pontuacao_maxima=Decimal('10.00'),
        )
        avaliacao = AvaliacaoQualidadeCompetencia.objects.create(competencia=competencia, modelo=modelo)
        AvaliacaoCriterioCompetencia.objects.create(avaliacao=avaliacao, criterio=criterio, nota_obtida=Decimal('8.00'))

        avaliacao.refresh_from_db()
        competencia.refresh_from_db()
        self.assertEqual(avaliacao.percentual_desempenho, Decimal('80.00'))
        self.assertEqual(competencia.valor_liberado, Decimal('80.00'))

    def test_reajuste_respeita_valor_referencial_e_calcula_retroatividade(self):
        item = ContratoItem.objects.create(
            contrato=self.contrato,
            ordem=1,
            descricao='Item reajustado',
            quantidade=Decimal('10.00'),
            valor_unitario=Decimal('90.00'),
            valor_referencial=Decimal('100.00'),
        )
        competencia = self.contrato.competencias.get(periodo_inicio=date(2026, 4, 1), periodo_fim=date(2026, 4, 30))
        MedicaoItemCompetencia.objects.create(
            competencia=competencia,
            item_contrato=item,
            quantidade=Decimal('2.00'),
            valor_unitario_aplicado=Decimal('90.00'),
        )
        evento = EventoFinanceiroContrato.objects.create(
            contrato=self.contrato,
            tipo=EventoFinanceiroContrato.Tipo.REAJUSTE,
            indice_aplicado='IPCA',
            data_base=date(2026, 4, 1),
            data_aplicacao=date(2026, 5, 1),
            percentual_aplicado=Decimal('10.00'),
        )
        EventoFinanceiroItem.objects.create(
            evento=evento,
            item_contrato=item,
            valor_original=Decimal('90.00'),
            valor_reajustado=Decimal('99.00'),
            valor_referencial=Decimal('100.00'),
        )

        memoria = evento.memorias.get(competencia=competencia, item_contrato=item)
        self.assertEqual(memoria.diferenca_total, Decimal('18.00'))


class ContratosViewTests(ContratosBaseTest):
    """Valida ACL, filtros de listagem e exportação do diário de bordo."""

    def test_acl_bloqueia_usuario_sem_regra(self):
        self.client.login(username='maria', password='123')
        response = self.client.get(reverse('contratos:home'))
        self.assertEqual(response.status_code, 403)

    def test_listagem_filtra_por_texto(self):
        self.client.login(username='joao', password='123')
        response = self.client.get(reverse('contratos:contrato_list'), {'q': 'Limpeza'})
        self.assertContains(response, 'Limpeza predial')

    def test_listagem_ordena_colunas_em_ordem_crescente(self):
        Contrato.objects.create(
            numero_contrato='02/2026',
            apelido='Zeladoria',
            objeto='Apoio operacional',
            detalhamento_objeto='Serviço complementar.',
            data_inicio_vigencia=date(2026, 2, 1),
            prazo_inicial_meses=12,
            vigencia_maxima_meses=24,
            empresa_contratada=self.empresa,
            fiscal_administrativo=self.admin,
            fiscal_tecnico=self.admin,
            gestor_contrato=self.admin,
            base_mensal=Decimal('2500.00'),
            valor_global=Decimal('30000.00'),
            criado_por=self.admin,
            atualizado_por=self.admin,
        )
        self.client.login(username='joao', password='123')

        response = self.client.get(reverse('contratos:contrato_list'), {'ordem': 'apelido', 'direcao': 'asc'})

        contratos = list(response.context['contratos'])
        self.assertEqual([contrato.apelido for contrato in contratos], ['Limpeza predial', 'Zeladoria'])

    def test_listagem_ordena_colunas_em_ordem_decrescente(self):
        Contrato.objects.create(
            numero_contrato='02/2026',
            apelido='Zeladoria',
            objeto='Apoio operacional',
            detalhamento_objeto='Serviço complementar.',
            data_inicio_vigencia=date(2026, 2, 1),
            prazo_inicial_meses=12,
            vigencia_maxima_meses=24,
            empresa_contratada=self.empresa,
            fiscal_administrativo=self.admin,
            fiscal_tecnico=self.admin,
            gestor_contrato=self.admin,
            base_mensal=Decimal('2500.00'),
            valor_global=Decimal('30000.00'),
            criado_por=self.admin,
            atualizado_por=self.admin,
        )
        self.client.login(username='joao', password='123')

        response = self.client.get(reverse('contratos:contrato_list'), {'ordem': 'apelido', 'direcao': 'desc'})

        contratos = list(response.context['contratos'])
        self.assertEqual([contrato.apelido for contrato in contratos], ['Zeladoria', 'Limpeza predial'])

    def test_detalhe_ordena_competencias_com_pagamentos_ao_final(self):
        ChecklistPagamentoModelo.objects.create(contrato=self.contrato, ordem=1, titulo='NF', obrigatorio=True)
        competencia_janeiro = self.contrato.competencias.get(periodo_inicio=date(2026, 1, 1))
        competencia_fevereiro = self.contrato.competencias.get(periodo_inicio=date(2026, 2, 1))
        competencia_marco = self.contrato.competencias.get(periodo_inicio=date(2026, 3, 1))
        competencia_fevereiro.status = CompetenciaPagamento.Status.PAGO
        competencia_fevereiro.save(update_fields=['status'])
        competencia_marco.status = CompetenciaPagamento.Status.EM_CONFERENCIA
        competencia_marco.save(update_fields=['status'])

        self.client.login(username='joao', password='123')
        response = self.client.get(reverse('contratos:contrato_detail', args=[self.contrato.pk]))

        competencias = list(response.context['competencias'])
        posicoes = {competencia.pk: indice for indice, competencia in enumerate(competencias)}
        self.assertLess(posicoes[competencia_janeiro.pk], posicoes[competencia_marco.pk])
        self.assertLess(posicoes[competencia_marco.pk], posicoes[competencia_fevereiro.pk])

    def test_competencia_nao_exibe_avaliacao_qualidade_sem_modelo_ativo(self):
        self.client.login(username='admin-contratos', password='123')

        response = self.client.get(reverse('contratos:contrato_detail', args=[self.contrato.pk]))

        self.assertNotContains(response, 'Criar avaliação')

    def test_competencia_exibe_avaliacao_qualidade_para_fiscal_com_modelo_ativo(self):
        ModeloAvaliacaoQualidade.objects.create(
            contrato=self.contrato,
            nome='Modelo ativo',
            vigencia_inicio=date(2026, 1, 1),
            ativo=True,
        )
        self.client.login(username='admin-contratos', password='123')

        response = self.client.get(reverse('contratos:contrato_detail', args=[self.contrato.pk]))

        self.assertContains(response, 'Criar avaliação')

    def test_rota_de_avaliacao_bloqueia_usuario_que_nao_e_fiscal(self):
        ModeloAvaliacaoQualidade.objects.create(
            contrato=self.contrato,
            nome='Modelo ativo',
            vigencia_inicio=date(2026, 1, 1),
            ativo=True,
        )
        competencia = self.contrato.competencias.get(periodo_inicio=date(2026, 1, 1))
        self.client.login(username='joao', password='123')

        response = self.client.get(reverse('contratos:avaliacao_create', args=[competencia.pk]))

        self.assertRedirects(response, reverse('contratos:contrato_detail', args=[self.contrato.pk]), fetch_redirect_response=False)

    def test_exclusao_de_contrato_com_vinculos_protegidos_redireciona_sem_erro(self):
        ChecklistPagamentoModelo.objects.create(contrato=self.contrato, titulo='NF mensal', obrigatorio=True)
        modelo = ModeloAvaliacaoQualidade.objects.create(
            contrato=self.contrato,
            nome='Modelo ativo',
            vigencia_inicio=date(2026, 1, 1),
            ativo=True,
        )
        grupo = GrupoAvaliacaoQualidade.objects.create(modelo=modelo, ordem=1, nome='Grupo', peso=Decimal('1.00'))
        criterio = CriterioAvaliacaoQualidade.objects.create(
            grupo=grupo,
            ordem=1,
            nome='Critério',
            peso=Decimal('1.00'),
            pontuacao_maxima=Decimal('10.00'),
        )
        competencia = self.contrato.competencias.get(periodo_inicio=date(2026, 1, 1))
        avaliacao = AvaliacaoQualidadeCompetencia.objects.create(competencia=competencia, modelo=modelo)
        AvaliacaoCriterioCompetencia.objects.create(avaliacao=avaliacao, criterio=criterio, nota_obtida=Decimal('8.00'))
        RegraAcesso.objects.create(recurso=self.recurso, usuario=self.admin, nivel=RegraAcesso.NIVEL_MODIFICACAO)
        self.client.login(username='admin-contratos', password='123')

        response = self.client.post(reverse('contratos:contrato_delete', args=[self.contrato.pk]))

        self.assertRedirects(response, reverse('contratos:contrato_detail', args=[self.contrato.pk]), fetch_redirect_response=False)
        self.assertTrue(Contrato.objects.filter(pk=self.contrato.pk).exists())

    def test_checklist_padrao_usa_proxima_ordem_quando_campo_fica_em_branco(self):
        ChecklistPagamentoModelo.objects.create(contrato=self.contrato, ordem=1, titulo='FGTS', obrigatorio=True)
        self.client.login(username='joao', password='123')

        response = self.client.post(
            reverse('contratos:checklist_modelo_create', args=[self.contrato.pk]),
            {
                'titulo': 'teste paga checklist 02',
                'descricao': '',
                'obrigatorio': 'on',
            },
        )

        item = ChecklistPagamentoModelo.objects.get(titulo='teste paga checklist 02')
        self.assertRedirects(response, reverse('contratos:contrato_detail', args=[self.contrato.pk]), fetch_redirect_response=False)
        self.assertEqual(item.ordem, 2)

    def test_bloqueia_medicao_enquanto_competencia_aguarda_checklist_padrao(self):
        competencia = self.contrato.competencias.get(periodo_inicio=date(2026, 1, 1))
        self.client.login(username='joao', password='123')

        response = self.client.get(reverse('contratos:medicao_create', args=[competencia.pk]))

        self.assertRedirects(response, reverse('contratos:contrato_detail', args=[self.contrato.pk]), fetch_redirect_response=False)

    def test_tela_de_medicao_traz_itens_do_contrato_automaticamente(self):
        ChecklistPagamentoModelo.objects.create(contrato=self.contrato, titulo='NF mensal', obrigatorio=True)
        item = ContratoItem.objects.create(
            contrato=self.contrato,
            ordem=1,
            descricao='Item mensal automatico',
            quantidade=Decimal('10.00'),
            valor_unitario=Decimal('50.00'),
        )
        competencia = self.contrato.competencias.get(periodo_inicio=date(2026, 1, 1))
        self.client.login(username='joao', password='123')

        response = self.client.get(reverse('contratos:medicao_create', args=[competencia.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Item mensal automatico')
        self.assertIn(f'quantidade_{item.pk}', response.content.decode())

    def test_post_de_medicao_em_lote_grava_quantidades_dos_itens(self):
        ChecklistPagamentoModelo.objects.create(contrato=self.contrato, titulo='NF mensal', obrigatorio=True)
        item1 = ContratoItem.objects.create(
            contrato=self.contrato,
            ordem=1,
            descricao='Item 1',
            quantidade=Decimal('10.00'),
            valor_unitario=Decimal('50.00'),
        )
        item2 = ContratoItem.objects.create(
            contrato=self.contrato,
            ordem=2,
            descricao='Item 2',
            quantidade=Decimal('4.00'),
            valor_unitario=Decimal('100.00'),
        )
        competencia = self.contrato.competencias.get(periodo_inicio=date(2026, 1, 1))
        self.client.login(username='joao', password='123')

        response = self.client.post(
            reverse('contratos:medicao_create', args=[competencia.pk]),
            {
                f'quantidade_{item1.pk}': '2.00',
                f'quantidade_{item2.pk}': '1.00',
            },
        )

        self.assertRedirects(response, reverse('contratos:contrato_detail', args=[self.contrato.pk]), fetch_redirect_response=False)
        competencia.refresh_from_db()
        self.assertEqual(competencia.medicoes.count(), 2)
        self.assertEqual(competencia.medicoes.get(item_contrato=item1).valor_unitario_aplicado, Decimal('50.00'))
        self.assertEqual(competencia.medicoes.get(item_contrato=item1).quantidade, Decimal('2.00'))
        self.assertEqual(competencia.medicoes.get(item_contrato=item2).quantidade, Decimal('1.00'))

    def test_cadastro_de_checklist_padrao_libera_fluxo_da_competencia(self):
        competencia = self.contrato.competencias.get(periodo_inicio=date(2026, 1, 1))
        self.client.login(username='joao', password='123')

        response = self.client.post(
            reverse('contratos:checklist_modelo_create', args=[self.contrato.pk]),
            {
                'titulo': 'NF mensal',
                'descricao': 'Documento obrigatório',
                'obrigatorio': 'on',
            },
        )

        competencia.refresh_from_db()
        self.assertRedirects(response, reverse('contratos:contrato_detail', args=[self.contrato.pk]), fetch_redirect_response=False)
        self.assertEqual(competencia.checklist_itens.count(), 1)
        self.assertFalse(competencia.aguardando_checklist_padrao)
        self.assertEqual(competencia.status, CompetenciaPagamento.Status.RASCUNHO)

    def test_modal_ajax_de_anexo_do_checklist_aceita_apenas_arquivo(self):
        competencia = self.contrato.competencias.get(periodo_inicio=date(2026, 1, 1))
        ChecklistPagamentoModelo.objects.create(contrato=self.contrato, titulo='NF mensal', obrigatorio=True)
        item = competencia.checklist_itens.first()
        self.client.login(username='joao', password='123')

        response_get = self.client.get(
            f"{reverse('contratos:checklist_anexo_create', args=[item.pk])}?modal=1",
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertEqual(response_get.status_code, 200)
        self.assertContains(response_get, 'Salvar anexo')

        response = self.client.post(
            f"{reverse('contratos:checklist_anexo_create', args=[item.pk])}?modal=1",
            {'arquivo': SimpleUploadedFile('checklist.pdf', b'%PDF-1.4 checklist', content_type='application/pdf')},
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        payload = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(payload['success'])
        item.refresh_from_db()
        self.assertEqual(item.anexos.count(), 1)
        self.assertTrue(item.concluido)
        self.assertIsNotNone(item.validado_em)

    def test_limpar_anexo_do_checklist_remove_arquivo_e_retorna_item_para_pendente(self):
        competencia = self.contrato.competencias.get(periodo_inicio=date(2026, 1, 1))
        ChecklistPagamentoModelo.objects.create(contrato=self.contrato, titulo='NF mensal', obrigatorio=True)
        item = competencia.checklist_itens.first()
        item.anexos.create(arquivo=SimpleUploadedFile('checklist.pdf', b'%PDF-1.4 checklist', content_type='application/pdf'))
        self.client.login(username='joao', password='123')

        response = self.client.post(reverse('contratos:checklist_anexo_delete', args=[item.pk]))

        self.assertRedirects(response, reverse('contratos:contrato_detail', args=[self.contrato.pk]), fetch_redirect_response=False)
        item.refresh_from_db()
        self.assertEqual(item.anexos.count(), 0)
        self.assertFalse(item.concluido)
        self.assertIsNone(item.validado_em)

    def test_anexo_do_checklist_rejeita_arquivo_que_nao_seja_pdf(self):
        competencia = self.contrato.competencias.get(periodo_inicio=date(2026, 1, 1))
        ChecklistPagamentoModelo.objects.create(contrato=self.contrato, titulo='NF mensal', obrigatorio=True)
        item = competencia.checklist_itens.first()
        self.client.login(username='joao', password='123')

        response = self.client.post(
            f"{reverse('contratos:checklist_anexo_create', args=[item.pk])}?modal=1",
            {'arquivo': SimpleUploadedFile('checklist.txt', b'conteudo demonstrativo', content_type='text/plain')},
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Envie um arquivo PDF válido.')

    def test_modal_ajax_de_pagamento_anexa_documentos_e_marca_competencia_como_paga(self):
        competencia = self.contrato.competencias.get(periodo_inicio=date(2026, 1, 1))
        ChecklistPagamentoModelo.objects.create(contrato=self.contrato, titulo='NF mensal', obrigatorio=True)
        competencia.refresh_from_db()
        for item in competencia.checklist_itens.all():
            item.concluido = True
            item.validado_em = timezone.now()
            item.save(update_fields=['concluido', 'validado_em'])
        self.client.login(username='joao', password='123')

        response_get = self.client.get(
            f"{reverse('contratos:competencia_pagamento_executar', args=[competencia.pk])}?modal=1",
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertEqual(response_get.status_code, 200)
        self.assertContains(response_get, 'Concluir pagamento')

        response_post = self.client.post(
            f"{reverse('contratos:competencia_pagamento_executar', args=[competencia.pk])}?modal=1",
            {
                'anexo_nota_fiscal': SimpleUploadedFile('nf.pdf', b'%PDF-1.4 nf', content_type='application/pdf'),
                'anexo_atestado_realizacao': SimpleUploadedFile('atestado.pdf', b'%PDF-1.4 atestado', content_type='application/pdf'),
                'anexo_despacho_dof': SimpleUploadedFile('dof.pdf', b'%PDF-1.4 dof', content_type='application/pdf'),
            },
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        payload = json.loads(response_post.content)
        self.assertEqual(response_post.status_code, 200)
        self.assertTrue(payload['success'])
        competencia.refresh_from_db()
        self.assertEqual(competencia.status, CompetenciaPagamento.Status.PAGO)
        self.assertTrue(bool(competencia.anexo_nota_fiscal))
        self.assertTrue(bool(competencia.anexo_atestado_realizacao))
        self.assertTrue(bool(competencia.anexo_despacho_dof))

    def test_exporta_diario_xlsx(self):
        OcorrenciaContrato.objects.create(
            contrato=self.contrato,
            data_registro=date(2026, 1, 10),
            tipo_ocorrencia='Fiscalização',
            descricao='Vistoria executada.',
            usuario=self.admin,
        )
        self.client.login(username='joao', password='123')
        response = self.client.get(reverse('contratos:ocorrencia_export_xlsx', args=[self.contrato.pk]))
        workbook = load_workbook(BytesIO(response.content))
        ws = workbook.active
        self.assertEqual(response.status_code, 200)
        self.assertEqual(ws['A1'].value, 'Número do Contrato')
        self.assertEqual(ws['B1'].value, '01/2026')
        self.assertEqual(ws['A6'].value, 'Data')

    def test_cadastro_ajax_de_empresa_retorna_json_para_modal(self):
        self.client.login(username='joao', password='123')
        response = self.client.post(
            f"{reverse('contratos:empresa_create')}?modal=1",
            {
                'razao_social': 'Empresa Modal',
                'cnpj': '33.333.333/0001-33',
                'nome_fantasia': 'Modal Ltda',
                'logradouro': 'Rua A',
                'numero': '10',
                'complemento': '',
                'bairro': 'Centro',
                'cidade': 'São Paulo',
                'estado': 'SP',
                'cep': '01000-000',
            },
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        payload = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(payload['success'])
        self.assertEqual(payload['empresa']['label'], 'Empresa Modal')

    def test_formulario_gera_numero_incremental_no_mesmo_ano(self):
        self.client.login(username='joao', password='123')
        Contrato.objects.create(
            numero_contrato='070/2026',
            apelido='Outro contrato',
            objeto='Objeto auxiliar',
            detalhamento_objeto='Detalhamento.',
            data_inicio_vigencia=date(2026, 2, 1),
            prazo_inicial_meses=12,
            vigencia_maxima_meses=24,
            empresa_contratada=self.empresa,
            fiscal_administrativo=self.admin,
            fiscal_tecnico=self.admin,
            gestor_contrato=self.admin,
            base_mensal=Decimal('500.00'),
            criado_por=self.admin,
            atualizado_por=self.admin,
        )

        response = self.client.post(
            reverse('contratos:contrato_create'),
            {
                'numero_contrato': '',
                'numero_contrato_incremental': 'on',
                'apelido': 'Contrato incremental',
                'objeto': 'Objeto incremental',
                'detalhamento_objeto': 'Detalhamento.',
                'data_inicio_vigencia': '2026-06-01',
                'prazo_inicial_meses': '12',
                'vigencia_maxima_meses': '24',
                'empresa_contratada': str(self.empresa.pk),
                'fiscal_administrativo': str(self.admin.pk),
                'fiscal_tecnico': str(self.admin.pk),
                'gestor_contrato': str(self.admin.pk),
                'base_mensal': '1000.00',
                'situacao_forcada': '',
                'detalhamento-TOTAL_FORMS': '0',
                'detalhamento-INITIAL_FORMS': '0',
                'detalhamento-MIN_NUM_FORMS': '0',
                'detalhamento-MAX_NUM_FORMS': '1000',
            },
        )

        criado = Contrato.objects.get(apelido='Contrato incremental')
        self.assertRedirects(response, reverse('contratos:contrato_detail', args=[criado.pk]), fetch_redirect_response=False)
        self.assertEqual(criado.numero_contrato, '071/2026')

    def test_formulario_incremental_reinicia_sequencia_em_novo_ano(self):
        form = ContratoForm(
            data={
                'numero_contrato': '',
                'numero_contrato_incremental': 'on',
                'apelido': 'Virada de ano',
                'objeto': 'Objeto',
                'detalhamento_objeto': 'Detalhamento.',
                'data_inicio_vigencia': '2027-01-15',
                'prazo_inicial_meses': '12',
                'vigencia_maxima_meses': '24',
                'empresa_contratada': str(self.empresa.pk),
                'fiscal_administrativo': str(self.admin.pk),
                'fiscal_tecnico': str(self.admin.pk),
                'gestor_contrato': str(self.admin.pk),
                'base_mensal': '1000.00',
                'situacao_forcada': '',
            }
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data['numero_contrato'], '001/2027')

    def test_formulario_valida_formato_manual_do_numero(self):
        form = ContratoForm(
            data={
                'numero_contrato': '1/2026',
                'apelido': 'Manual inválido',
                'objeto': 'Objeto',
                'detalhamento_objeto': 'Detalhamento.',
                'data_inicio_vigencia': '2026-01-01',
                'prazo_inicial_meses': '12',
                'vigencia_maxima_meses': '24',
                'empresa_contratada': str(self.empresa.pk),
                'fiscal_administrativo': str(self.admin.pk),
                'fiscal_tecnico': str(self.admin.pk),
                'gestor_contrato': str(self.admin.pk),
                'base_mensal': '1000.00',
                'situacao_forcada': '',
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn('numero_contrato', form.errors)

    def test_criacao_de_contrato_salva_detalhamento_por_itens(self):
        self.client.login(username='joao', password='123')
        response = self.client.post(
            reverse('contratos:contrato_create'),
            {
                'numero_contrato': '',
                'numero_contrato_incremental': 'on',
                'apelido': 'Contrato com detalhamento',
                'objeto': 'Objeto detalhado',
                'data_inicio_vigencia': '2026-06-01',
                'prazo_inicial_meses': '12',
                'vigencia_maxima_meses': '24',
                'empresa_contratada': str(self.empresa.pk),
                'fiscal_administrativo': str(self.admin.pk),
                'fiscal_tecnico': str(self.admin.pk),
                'gestor_contrato': str(self.admin.pk),
                'base_mensal': '1000.00',
                'situacao_forcada': '',
                'detalhamento-TOTAL_FORMS': '2',
                'detalhamento-INITIAL_FORMS': '0',
                'detalhamento-MIN_NUM_FORMS': '0',
                'detalhamento-MAX_NUM_FORMS': '1000',
                'detalhamento-0-ordem': '1',
                'detalhamento-0-descricao': 'Primeiro item do detalhamento',
                'detalhamento-1-ordem': '2',
                'detalhamento-1-descricao': 'Segundo item do detalhamento',
            },
        )

        contrato = Contrato.objects.get(apelido='Contrato com detalhamento')
        self.assertRedirects(response, reverse('contratos:contrato_detail', args=[contrato.pk]), fetch_redirect_response=False)
        self.assertEqual(contrato.detalhamento_itens.count(), 2)
        self.assertEqual(
            list(contrato.detalhamento_itens.order_by('ordem').values_list('descricao', flat=True)),
            ['Primeiro item do detalhamento', 'Segundo item do detalhamento'],
        )
        self.assertEqual(
            contrato.detalhamento_objeto,
            '1. Primeiro item do detalhamento\n2. Segundo item do detalhamento',
        )
